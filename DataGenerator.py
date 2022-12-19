import psycopg2
import random
import openpyxl
from pandas import *

# populating global college campus data that will not change throughout the simulation
def populateGlobalData(conn):
    #static
    addClubs(conn)
    addCourseAtt(conn)
    addCourseInfo(conn)
    addCoursePrereq(conn)
    addDepartments(conn)
    addHallInfo(conn)
    addHallRoom(conn)
    addJobs(conn)
    addLibraries(conn)
    addLibraryItems(conn)
    addMajor(conn)
    addMajorAttReq(conn)
    addMajorCourseReq(conn)
    addResLifeStaff(conn)
    addStaffProfile(conn)

def initializeSemesterStart(semester, year, addition, conn):    
    addStudentProfile(conn, addition)
    addStudentJobs(conn, addition)
    addClubMembers(conn, addition)
    addEvents(semester, year, conn)
    makeLibraryService(conn)
    makeLibraryItemLoan(conn,year)
    

# populating clubs table
def addClubs(connection):
    cur = connection.cursor()

    with open('Data/Clubs.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'clubs', sep=',')

    connection.commit()

# populating course attributes table


def addCourseAtt(connection):
    cur = connection.cursor()

    with open('Data/CourseAtt.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'course_attributes', sep=',')

    connection.commit()

# populating course information table


def addCourseInfo(connection):
    cur = connection.cursor()

    with open('Data/CourseProfile.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'course_profile', sep=',')

    connection.commit()

# populating course prerequisites table


def addCoursePrereq(connection):
    cur = connection.cursor()

    with open('Data/CoursePrereq.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'course_prerequisites', sep=',')

    connection.commit()

# populating departments table


def addDepartments(connection):
    cur = connection.cursor()

    with open('Data/Departments.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'departments', sep=',')

    connection.commit()

# populating residence hall information table


def addHallInfo(connection):
    cur = connection.cursor()

    with open('Data/HallInfo.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'hall_info', sep=',')

    connection.commit()

# populating residence hall rooms table


def addHallRoom(connection):
    cur = connection.cursor()

    with open('Data/HallRoom.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'hall_room', sep=',')

    connection.commit()

# populating jobs table


def addJobs(connection):
    cur = connection.cursor()

    with open('Data/Jobs.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'jobs', sep=',')

    connection.commit()

# populating libraries table


def addLibraries(connection):
    cur = connection.cursor()

    with open('Data/Libraries.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'libraries', sep=',')

    connection.commit()

# populating library items table


def addLibraryItems(connection):
    cur = connection.cursor()

    with open('Data/LibraryItems.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'library_items', sep=',')

    connection.commit()

# populating major table


def addMajor(connection):
    cur = connection.cursor()

    with open('Data/Major.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'major', sep=',')

    connection.commit()

# populating major attribute requirement table


def addMajorAttReq(connection):
    cur = connection.cursor()

    with open('Data/MajorAttReq.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'major_attribute_requirement', sep=',')

    connection.commit()

# populating major course requirement table


def addMajorCourseReq(connection):
    cur = connection.cursor()

    with open('Data/MajorCourseReq.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'major_course_requirement', sep=',')

    connection.commit()

# populating residence life staff table


def addResLifeStaff(connection):
    cur = connection.cursor()

    with open('Data/ResLifeStaff.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'reslife_staff', sep=',')

    connection.commit()

# populating staff profile table


def addStaffProfile(connection):
    cur = connection.cursor()

    with open('Data/StaffProfile.csv', 'r') as f:
        next(f)  # Skip the header row.
        cur.copy_from(f, 'staff_profile', sep=',')

    connection.commit()


# ||||||||||||||||||||||||||||||||||||
# NON-STATIC DATA
# ||||||||||||||||||||||||||||||||||||


def initializeSemesterEnd(semester, year, conn):
    makeStudentsTakeCourses(semester, year, conn)
    giveStudentsGrades(conn)
    addStudentSemester(semester, year, conn)


# populating student profile table
def addStudentProfile(connection, addition):
    file = read_csv('Data/StudentProfile.csv')
    cur = connection.cursor()

    for i in range(addition-100, addition):
        values = (
            file['L_number'][i].astype(Int64Dtype), 
            file['Department'][i], 
            file['first_name'][i], 
            file['last_name'][i], 
            file['DOB'][i],
            file['Major'][i],
            file['Email'][i],
            file['Year'][i].astype(Int64Dtype),
            file['Phone_#'][i],
            file['Emergency_Phone_#'][i]
            )

        cur.execute("INSERT INTO student_profile VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", values)
        connection.commit()
    

# populating student club members
def addClubMembers(connection, addition):
    cur = connection.cursor()
    connection.autocommit = True

    cur.execute("SELECT lnumber from student_profile")
    studentsList = cur.fetchall()
    studentsList = [x[0] for x in studentsList]

    cur.execute("SELECT clubid from clubs")
    clubList = cur.fetchall()
    clubList = [x[0] for x in clubList]

    general = read_csv("Data/General.csv")
    memberTitles = general['Member Title'].tolist()
    memberTitles = [x for x in memberTitles if str(x) != 'nan']

    for i in range(addition-100, addition):
        currentStudent = studentsList[i]
        randomClub = random.randint(0, len(clubList))
        randomMemberTitle = random.randint(0, len(memberTitles))

        currentClub = clubList[randomClub-1]
        currentMemberTitle = memberTitles[randomMemberTitle-1]

        query = """INSERT INTO club_members VALUES (%s,%s,%s)"""
        insertValues = (currentStudent, currentClub, currentMemberTitle)
        cur.execute(query, insertValues)

# populating student jobs


def addStudentJobs(connection, addition):
    cur = connection.cursor()
    connection.autocommit = True

    cur.execute("SELECT lnumber from student_profile")
    studentsList = cur.fetchall()
    studentsList = [x[0] for x in studentsList]

    cur.execute("SELECT jobname, deptname from jobs")
    jobsList = cur.fetchall()
    jobsList = [x[0] for x in jobsList]

    cur.execute("SELECT deptname from jobs")
    jobDeptNameList = cur.fetchall()
    jobDeptNameList = [x[0] for x in jobDeptNameList]

    for i in range(addition-100, addition):
        currentStudent = studentsList[i]
        randomJob = random.randint(0, len(jobsList))

        currentJob = jobsList[randomJob-1]
        currentJobDeptName = jobDeptNameList[randomJob-1]

        query = """INSERT INTO student_job VALUES (%s,%s,%s)"""
        insertValues = (currentStudent, currentJob, currentJobDeptName)
        cur.execute(query, insertValues)

# adding events


def addEvents(semester, year, connection):
    cur = connection.cursor()
    connection.autocommit = True

    events = read_csv("Data/Events.csv")

    eventNames = events['Event_Name'].tolist()
    eventNames = [x for x in eventNames if str(x) != 'nan']

    eventLocations = events['Location'].tolist()
    eventLocations = [x for x in eventLocations if str(x) != 'nan']

    cur.execute("SELECT clubid from clubs")
    clubList = cur.fetchall()
    clubList = [x[0] for x in clubList]

    clubsWithEventsAmount = random.randint(1, len(clubList) - 1)
    count = 1

    while (count < clubsWithEventsAmount):
        randomName = random.randint(0, len(eventNames))
        randomLocation = random.randint(0, len(eventLocations))
        randomClub = random.randint(0, len(clubList))

        currentClub = clubList[randomClub-1]
        currentEvent = currentClub * random.randint(1,10000)
        currentName = eventNames[randomName-1]
        currentLocation = eventLocations[randomLocation-1]

        if(semester == "Fall"):  
            randomMonth = random.randint(8, 12)
        else:
            randomMonth = random.randint(1, 5)
        randomDay = random.randint(1, 28)
        currentDate = f"{randomMonth}/{randomDay}/{year}"

        query = """INSERT INTO events VALUES (%s,%s,%s,%s,%s)"""
        insertValues = (currentEvent, currentName,
                        currentLocation, currentDate, currentClub)
        cur.execute(query, insertValues)

        count += 1


def addStudentSemester(semester, year, connection):
    cur = connection.cursor()
    connection.autocommit = True

    cur.execute("SELECT lnumber from student_profile")
    studentsList = cur.fetchall()
    studentsList = [x[0] for x in studentsList]

    cur.execute("SELECT hallname from hall_room")
    hallsList = cur.fetchall()
    hallsList = [x[0] for x in hallsList]

    cur.execute("SELECT roomnumber from hall_room")
    roomsList = cur.fetchall()
    roomsList = [x[0] for x in roomsList]

    general = read_csv("Data/General.csv")

    financialTypes = general['Financial Type'].tolist()
    financialTypes = [x for x in financialTypes if str(x) != 'nan']

    enrolStatuses = general['Enrollment Status'].tolist()
    enrolStatuses = [x for x in enrolStatuses if str(x) != 'nan']

    mealPlans = general['Meal Plan Per Week'].tolist()
    mealPlans = [x for x in mealPlans if str(x) != 'nan']

    count = 1
    while (count < len(studentsList)):
        currentStudent = studentsList[count]

        randomFinancialType = random.randint(0, len(financialTypes))
        randomEnrolStatus = random.randint(0, len(enrolStatuses))
        randomMealPlan = random.randint(0, len(mealPlans))
        randomHall = random.randint(0, len(hallsList))

        currentFinancialType = financialTypes[randomFinancialType-1]
        currentEnrolStatus = enrolStatuses[randomEnrolStatus-1]
        currentMealPlan = mealPlans[randomMealPlan-1]
        currentHall = hallsList[randomHall-1]
        currentRoom = roomsList[randomHall-1]

        cur.execute("SELECT grade FROM grade_report WHERE lnumber = %s", [
                    currentStudent])
        allGrades = cur.fetchall()
        allGrades = [x[0] for x in allGrades]
        gradesSum = sum(float(i) for i in allGrades)
        termGrade = gradesSum / len(allGrades)

        query = """INSERT INTO student_semester VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
        insertValues = (
            currentStudent,
            semester,
            year,
            currentFinancialType,
            currentEnrolStatus,
            currentMealPlan,
            termGrade,
            currentHall,
            currentRoom
        )
        cur.execute(query, insertValues)

        count += 1


def makeStudentsTakeCourses(semester, year, conn):
    cur = conn.cursor()

    gettingStudentIds = "Select Major, lNumber FROM student_profile"

    cur.execute(gettingStudentIds)
    #print("Selecting rows from mobile table using cursor.fetchall")
    infos = cur.fetchall()

    for info in infos:
        # #print(info)
        major = info[0]
        id = info[1]
        studentCourses = "Select sectNum FROM student_courses WHERE lNumber=" + \
            str(id)
        cur.execute(studentCourses)
        sections = cur.fetchall()
        # if len(sections) > 0:
        #     print(sections[len(sections)-1][0])

        if len(sections) == 0:
            if major == "CS":
                csClasses = "Select courseSect FROM course_profile WHERE courseCRN=1001"
                cur.execute(csClasses)
                sections = cur.fetchall()
                max = len(sections)
                pickSection = random.randint(0, max-1)
                sectionNumber = sections[pickSection][0]
                fstCourse = sectionNumber
                # #print(fstCourse)
                csClasses = "Select courseSect FROM course_profile WHERE courseCRN=1002"
                cur.execute(csClasses)
                sections = cur.fetchall()
                max = len(sections)
                pickSection = random.randint(0, max-1)
                sectionNumber = sections[pickSection][0]
                sndCourse = sectionNumber
                # #print(sndCourse)
                insertCourse1 = "INSERT INTO student_courses Values("+str(
                    id)+",'"+semester+"',"+str(year)+",1001,"+str(fstCourse)+")"
                insertCourse2 = "INSERT INTO student_courses Values("+str(
                    id)+",'"+semester+"',"+str(year)+",1002,"+str(sndCourse)+")"
                # print(insertCourse1)
                # print(insertCourse2)
                cur.execute(insertCourse1)
                conn.commit()
                cur.execute(insertCourse2)
                conn.commit()
            if major == "ECON":
                csClasses = "Select courseSect FROM course_profile WHERE courseCRN=2001"
                cur.execute(csClasses)
                sections = cur.fetchall()
                max = len(sections)
                pickSection = random.randint(0, max-1)
                sectionNumber = sections[pickSection][0]
                fstCourse = sectionNumber
                # print(fstCourse)
                csClasses = "Select courseSect FROM course_profile WHERE courseCRN=2002"
                cur.execute(csClasses)
                sections = cur.fetchall()
                max = len(sections)
                pickSection = random.randint(0, max-1)
                sectionNumber = sections[pickSection][0]
                sndCourse = sectionNumber
                # print(sndCourse)
                insertCourse1 = "INSERT INTO student_courses Values("+str(
                    id)+",'"+semester+"',"+str(year)+",2001,"+str(fstCourse)+")"
                insertCourse2 = "INSERT INTO student_courses Values("+str(
                    id)+",'"+semester+"',"+str(year)+",2002,"+str(sndCourse)+")"
                # print(insertCourse1)
                # print(insertCourse2)
                cur.execute(insertCourse1)
                conn.commit()
                cur.execute(insertCourse2)
                conn.commit()
            if major == "MATH":
                csClasses = "Select courseSect FROM course_profile WHERE courseCRN=3001"
                cur.execute(csClasses)
                sections = cur.fetchall()
                max = len(sections)
                pickSection = random.randint(0, max-1)
                sectionNumber = sections[pickSection][0]
                fstCourse = sectionNumber
                # print(fstCourse)
                csClasses = "Select courseSect FROM course_profile WHERE courseCRN=3002"
                cur.execute(csClasses)
                sections = cur.fetchall()
                max = len(sections)
                pickSection = random.randint(0, max-1)
                sectionNumber = sections[pickSection][0]
                sndCourse = sectionNumber
                # print(sndCourse)
                insertCourse1 = "INSERT INTO student_courses Values("+str(
                    id)+",'"+semester+"',"+str(year)+",3001,"+str(fstCourse)+")"
                insertCourse2 = "INSERT INTO student_courses Values("+str(
                    id)+",'"+semester+"',"+str(year)+",3002,"+str(sndCourse)+")"
                # print(insertCourse1)
                # print(insertCourse2)
                cur.execute(insertCourse1)
                conn.commit()
                cur.execute(insertCourse2)
                conn.commit()
        else:
            lastSection = sections[len(sections)-1][0]
            last_digit = int(repr(lastSection)[-2])
            isOne = int(repr(lastSection)[-3])
            if last_digit+2 > 6 and isOne == 1:
                continue
                #print("the student have graduated")
            else:
                csClasses = "Select courseCRN FROM student_courses WHERE sectNum=" + \
                    str(lastSection)
                cur.execute(csClasses)
                courses = cur.fetchall()
                course = courses[0][0]
                course1 = course+1
                course2 = course+2
                csClasses = "Select courseSect FROM course_profile WHERE courseCRN=" + \
                    str(course1)
                cur.execute(csClasses)
                sections = cur.fetchall()
                max = len(sections)
                pickSection = random.randint(0, max-1)
                sectionNumber = sections[pickSection][0]
                fstCourse = sectionNumber
                # print(fstCourse)
                csClasses = "Select courseSect FROM course_profile WHERE courseCRN=" + \
                    str(course2)
                cur.execute(csClasses)
                sections = cur.fetchall()
                max = len(sections)
                pickSection = random.randint(0, max-1)
                sectionNumber = sections[pickSection][0]
                sndCourse = sectionNumber
                # print("here:"+str(sndCourse))
                insertCourse1 = "INSERT INTO student_courses Values("+str(
                    id)+",'"+semester+"',"+str(year)+","+str(course1)+","+str(fstCourse)+")"
                insertCourse2 = "INSERT INTO student_courses Values("+str(
                    id)+",'"+semester+"',"+str(year)+","+str(course2)+","+str(sndCourse)+")"
                # print(insertCourse1)
                # print(insertCourse2)
                cur.execute(insertCourse1)
                conn.commit()
                cur.execute(insertCourse2)
                conn.commit()


def giveStudentsGrades(conn):
    cur = conn.cursor()
    gettingStudentIds = "Select Major, lNumber FROM student_profile"

    cur.execute(gettingStudentIds)
    #print("Selecting rows from mobile table using cursor.fetchall")
    infos = cur.fetchall()

    for info in infos:
        # print(info)
        major = info[0]
        id = info[1]
        studentCourses = "Select sectNum FROM student_courses WHERE lNumber=" + \
            str(id)
        cur.execute(studentCourses)
        sections = cur.fetchall()
        beforeLastCourse = sections[len(sections)-2][0]
        lastCourse = sections[len(sections)-1][0]

        grade1 = general("Grade")
        grade2 = general("Grade")

        insertGrade1 = "INSERT INTO grade_report Values("+str(
            id)+","+str(beforeLastCourse)+","+str(grade1)+")"
        insertGrade2 = "INSERT INTO grade_report Values("+str(
            id)+","+str(lastCourse)+","+str(grade2)+")"
        # print(insertGrade1)
        # print(insertGrade2)
        cur.execute(insertGrade1)
        conn.commit()
        cur.execute(insertGrade2)
        conn.commit()

def makeLibraryItemLoan(conn,year):
    cur = conn.cursor()
    gettingStudentIds = "Select Major, lNumber FROM student_profile"
    cur.execute(gettingStudentIds)
    infos = cur.fetchall()
    getLibraryIds = "Select libraryID FROM libraries"
    getIteamIds = "Select libraryID FROM libraries"
    cur.execute(getLibraryIds)
    #print("Selecting rows from mobile table using cursor.fetchall")
    libList = cur.fetchall()
    cur.execute(getIteamIds)
    itemList = cur.fetchall()
    for i in range(0,11):
        pickStudent = random.randint(0,len(infos)-1)
        pickIteam = random.randint(0,len(itemList)-1)
        pickLibrary = random.randint(0,1)
        studentID = infos[pickStudent][1]
        libraryID = libList[pickLibrary][0]
        itemID = itemList[pickIteam][0]
        yearOut = random.randint(2010,year)
        dayOut = random.randint(1,30)
        monthOut = random.randint(1,12)
        dateOut = str(monthOut)+"/"+str(dayOut)+"/"+str(yearOut)
        # print("dateOut: "+dateOut)
        yearDue = random.randint(yearOut,year)
        dayDue = random.randint(1,30)
        monthDue = random.randint(1,12)
        dateDue = str(monthDue)+"/"+str(dayDue)+"/"+str(yearDue)
        # print("dateDue: "+dateDue)
        yearIn = random.randint(yearOut,yearDue)
        dayIn = random.randint(1,30)
        monthIn = random.randint(1,12)
        dateIn = str(monthIn)+"/"+str(dayIn)+"/"+str(yearIn)
        # print("dateIn: "+dateIn)
        insertLoan = "INSERT INTO library_loans Values(DEFAULT,"+str(itemID)+","+str(libraryID)+","+str(studentID)+",'"+str(dateOut)+"','"+str(dateIn)+"','"+str(dateDue)+"')"
        # print(insertLoan)
        cur.execute(insertLoan)
        conn.commit()

def makeLibraryService(conn):
    cur = conn.cursor()
    gettingStudentIds = "Select Major, lNumber FROM student_profile"
    cur.execute(gettingStudentIds)
    infos = cur.fetchall()
    
    #make Library service
    getLibraryIds = "Select libraryID FROM libraries"
    cur.execute(getLibraryIds)
    #print("Selecting rows from mobile table using cursor.fetchall")
    libList = cur.fetchall()

    for i in range(0,11):
        pickStudent = random.randint(0,len(infos))
        pickLibrary = random.randint(0,1)
        # print(pickStudent)
        studentID = infos[pickStudent][1]
        libraryID = libList[pickLibrary][0]
        # print(studentID)
        # print(libraryID)
        serviceName = general("Service Name")
        insertCourse1 = "INSERT INTO library_services Values(DEFAULT,"+str(studentID)+",'"+serviceName+"',"+str(libraryID)+")"
        # print(insertCourse1)
        cur.execute(insertCourse1)
        conn.commit()
        
def general(colName):
    path = "D:\CS\CollegeSimulationPostgres\Data\General.xlsx"

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj["Sheet1"]

    generalCols = {}
    col = 1
    row = 1
    while (sheet_obj.cell(column=col, row=1).value != None):
        name = sheet_obj.cell(column=col, row=1).value
        generalCols[name] = col
        col = col + 1

    colist = ["A", "B", "C", "D", "E", "F", "G", "H", "I", ]

    colNum = generalCols[colName]
    colIndex = colNum - 1

    colLength = 0
    for cell in sheet_obj[colist[colIndex]]:
        if cell.value == None:
            break
        colLength = colLength + 1

    row = random.randint(2, colLength)
    finalValue = sheet_obj.cell(column=colNum, row=row).value

    return finalValue
